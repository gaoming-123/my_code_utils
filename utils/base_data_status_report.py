# -*- coding: UTF-8 -*-
# Date   : 2020/3/9 11:03
# Editor : gmj
# Desc   : 统计全部数据情况
import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from common.database.mysql import MysqlConnect
from common.database.db_config import ALI_MYSQL_CONFIG

mysql_cnn = MysqlConnect(ALI_MYSQL_CONFIG)

QUERY_DICT = {
    't_major': {
        '专业数量': 'SELECT count(*) from t_major_operate',
    },

}


def report_base_data_status():
    file_path = f'./{datetime.date.today()}数据统计报告.xlsx'
    if os.path.exists(file_path):
        os.remove(file_path)
    wb = Workbook()
    for table, value in QUERY_DICT.items():
        print(table)
        eval(f'report_{table}(wb,"{file_path}",{value})')


# 院校相关
def report_t_school(wb, save_path, query_dict):
    # wb = Workbook()
    style = PatternFill("solid", fgColor="E2EFDA")
    ws = wb.create_sheet(index=0, title="学校信息报告")
    ws.column_dimensions['A'].width = 23.0
    ws['A1'] = '学校相关数据缺失统计结果'
    ws['A2'] = '学校数量'
    total = int(mysql_cnn.count('SELECT count(*) from (SELECT distinct  school_name from t_school_operate ) aa'))
    ws['B2'] = total
    # print(total)
    query_every(query_dict, save_path, total, wb, ws)


def query_every(query_dict, save_path, total, wb, ws):
    ws['A4'] = '数据字段'
    ws['B4'] = '数据总计'
    ws['C4'] = '现有数量'
    ws['D4'] = '缺失数量'
    row = 5
    style = PatternFill("solid", fgColor="E2EFDA")
    style1 = PatternFill("solid", fgColor="66CD00")
    style2 = PatternFill("solid", fgColor="EEEE00")
    for k, sql_txt in query_dict.items():
        row += 1
        ws.cell(row, 1).value = k
        ws.cell(row, 1).fill = style
        ws.cell(row, 2).value = total
        num = int(mysql_cnn.count(sql_txt))
        ws.cell(row, 3).value = num
        ws.cell(row, 3).fill = style1
        ws.cell(row, 4).value = total - num
        ws.cell(row, 4).fill = style2
    wb.save(save_path)


# 专业相关
def report_t_major(wb, save_path, query_dict):
    # wb = Workbook()
    style = PatternFill("solid", fgColor="E2EFDA")
    ws = wb.create_sheet(index=1, title="专业信息报告")
    ws.column_dimensions['A'].width = 23.0
    ws['A1'] = '专业相关数据缺失统计结果'
    ws['A2'] = '专业数量'
    total = int(mysql_cnn.count('SELECT count(*) from t_major_operate'))
    ws['B2'] = total
    # print(total)
    query_every(query_dict, save_path, total, wb, ws)


# 职业相关
def report_t_job(wb, save_path, query_dict):
    # wb = Workbook()
    style = PatternFill("solid", fgColor="E2EFDA")
    ws = wb.create_sheet(index=2, title="职业数据报告")
    ws.column_dimensions['A'].width = 20.0
    ws['A1'] = '职业数据缺失统计结果'
    ws['A2'] = '一级类数量'
    ws['B2'] = mysql_cnn.count('SELECT count(*) from t_job_first_type_operate')
    ws['A3'] = '二级类数量'
    ws['B3'] = mysql_cnn.count('SELECT count(*) from t_job_type_operate')
    ws['A4'] = '职位数量'
    total = int(mysql_cnn.count('SELECT count(*) from t_job_operate'))
    ws['B4'] = total

    ws['A4'] = '数据字段'
    ws['B4'] = '数据总计'
    ws['C4'] = '现有数量'
    ws['D4'] = '缺失数量'
    row = 5
    style = PatternFill("solid", fgColor="E2EFDA")
    style1 = PatternFill("solid", fgColor="66CD00")
    style2 = PatternFill("solid", fgColor="EEEE00")
    for k, sql_txt in query_dict.items():
        row += 1
        ws.cell(row, 1).value = k
        ws.cell(row, 1).fill = style
        ws.cell(row, 2).value = total
        num = int(mysql_cnn.count(sql_txt))
        ws.cell(row, 3).value = num
        ws.cell(row, 3).fill = style1
        ws.cell(row, 4).value = total - num
        ws.cell(row, 4).fill = style2
    wb.save(save_path)


# 省控线，一分一段表
def report_batch_one_point(wb, save_path, query_dict):
    # wb = Workbook()
    style = PatternFill("solid", fgColor="E2EFDA")
    ws = wb.create_sheet(index=3, title="一分一段，同分去向报告")
    ws.column_dimensions['A'].width = 20.0
    ws['A1'] = '一分一段，同分去向，省控线 相关数据缺失统计结果'
    ws['A2'] = '省份数量'
    total = 31
    ws['B2'] = total
    # print(total)
    query_every(query_dict, save_path, total, wb, ws)


MISS_DICT = {
    '职业相关数据缺失统计': {
        '学历分布缺失': "SELECT job_name from t_job_operate where job_id not in (SELECT DISTINCT job_id from t_job_education_operate)",   
    },
}


def output_data_missing():
    for k, output_items in MISS_DICT.items():
        file_path = f'./{datetime.date.today()}{k}.xlsx'
        if os.path.exists(file_path):
            os.remove(file_path)
        wb = Workbook()
        sheet_index = 0
        for theme, sql_txt in output_items.items():
            datas = mysql_cnn.fetchall(sql_txt)
            if datas:
                ws = wb.create_sheet(index=sheet_index, title=theme)
                for row_num, da in enumerate(datas):
                    for col_num, v in enumerate(list(da.values())):
                        ws.cell(row_num + 2, col_num + 1).value = v
                sheet_index += 1
        wb.save(file_path)


if __name__ == '__main__':
    # report_base_data_status()
    output_data_missing()
