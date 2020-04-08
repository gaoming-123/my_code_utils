# -*- coding: UTF-8 -*-
# Date   : 2020/3/30 16:29
# Editor : gmj
# Desc   : 
import openpyxl
# import xlwt  用于写xls文件
from xlrd import open_workbook  # xlrd用于读取xld


# 读取excel文件的公用方法
def read_xlsx(file_path, table_name):
    workbook = openpyxl.load_workbook(file_path)
    # shenames = workbook.get_sheet_names()
    worksheet = workbook.get_sheet_by_name(table_name)

    '''对sheet对象进行操作'''
    # name = worksheet.title
    # rows = worksheet.max_row
    # columns = worksheet.max_column
    result = []
    for row in worksheet.rows:
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        result.append(tuple(row_data))
        # print(row[0].value, row[1].value)
    return result


def read_xls(file_path, table_name):
    import xlwt  # 用于写入xls
    workbook = open_workbook(file_path)  # 打开xls文件
    # sheet_name = workbook.sheet_names()  # 打印所有sheet名称，是个列表
    # sheet = workbook.sheet_by_index(0)  # 根据sheet索引读取sheet中的所有内容
    worksheet = workbook.sheet_by_name(table_name)  # 根据sheet名称读取sheet中的所有内容
    result = []
    for i in range(worksheet.nrows):
        result.append(tuple(worksheet.row_values(i)))
    return result


# 读取csv的公用方法
def read_csv(file_path):
    files = list(open(file_path))
    result = []
    for line in files:
        result.append(tuple(line.my_strip().split(',')))
    return result


# 读取TXT的公用方法
def read_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as fr:
        content = fr.read()
    return content


if __name__ == '__main__':
    data = read_csv('./yifenyiduan上海.txt')
    print(data)
